"""
日報メール自動作成スクリプト

日報Excelから数値を集計し、Gmailスレッドに返信形式の下書きを作成する。
送信はしない。

Usage:
    python nippou_mail.py              # 自動判定（14:30-24:00→当日、0:00-14:30→前日）
    python nippou_mail.py 2026-04-17   # 日付指定
"""

import base64
import io
import json
import os
import re
import sys
from datetime import date, datetime, time, timedelta
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.header import Header
from pathlib import Path
from urllib.parse import quote

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

# --- 設定（環境変数またはここを書き換え） ---
USER_EMAIL = os.environ.get("NIPPOU_USER_EMAIL", "")
NIPPOU_TO = os.environ.get("NIPPOU_TO", "nippou@nksysg.com")
TOKEN_PATH = Path(os.environ.get("NIPPOU_TOKEN_PATH", str(Path.home() / ".google_workspace_mcp" / "credentials" / f"{USER_EMAIL}.json")))
NIPPOU_DIR = Path(os.environ.get("NIPPOU_DIR", ""))
NIPPOU_EXCEL_TEMPLATE = os.environ.get("NIPPOU_EXCEL_TEMPLATE", "日報{year}年{month}月【{classroom}】.xlsx")
CLASSROOM_NAME = os.environ.get("NIPPOU_CLASSROOM_NAME", "")
SEITOBETSU_DIR = NIPPOU_DIR / "生徒別一覧"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def get_target_date() -> date:
    """対象日を決定。引数指定 or 時刻で自動判定"""
    if len(sys.argv) >= 2:
        return datetime.strptime(sys.argv[1], "%Y-%m-%d").date()
    now = datetime.now()
    if now.time() < time(14, 30):
        return (now - timedelta(days=1)).date()
    return now.date()


def get_gmail_service():
    """Gmail APIサービスを取得"""
    with open(TOKEN_PATH) as f:
        token_data = json.load(f)
    creds = Credentials(
        token=token_data["token"],
        refresh_token=token_data["refresh_token"],
        token_uri=token_data["token_uri"],
        client_id=token_data["client_id"],
        client_secret=token_data["client_secret"],
        scopes=token_data["scopes"],
    )
    if creds.expired:
        creds.refresh(Request())
        # トークンを更新保存
        token_data["token"] = creds.token
        token_data["expiry"] = creds.expiry.isoformat() if creds.expiry else None
        with open(TOKEN_PATH, "w") as f:
            json.dump(token_data, f)
    return build("gmail", "v1", credentials=creds)


def count_from_excel(target: date) -> dict:
    """日報Excelから問合せ・入会・退会を集計"""
    year, month = target.year, target.month
    filename = NIPPOU_EXCEL_TEMPLATE.format(year=year, month=month)
    path = NIPPOU_DIR / "日報" / filename

    if not path.exists():
        raise FileNotFoundError(f"日報Excelが見つかりません: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)

    # --- 問合リスト ---
    ws = wb["問合リスト"]
    month_start = date(year, month, 1)
    toiawase_today = 0
    toiawase_month = 0
    nyukai_today = 0
    nyukai_month = 0

    for row in range(5, 500):
        b = ws.cell(row=row, column=2).value  # B: 日付
        if b is None:
            break
        d = b.date() if hasattr(b, "date") else b if isinstance(b, date) else None
        if d is None:
            continue

        if d == target:
            toiawase_today += 1
        if month_start <= d <= target:
            toiawase_month += 1

        # 入会: Q列=入塾手続日, X列=ステータス
        q = ws.cell(row=row, column=17).value
        x = ws.cell(row=row, column=24).value
        if q is not None and x is not None and "入会" in str(x):
            qd = q.date() if hasattr(q, "date") else q if isinstance(q, date) else None
            if qd is None:
                continue
            if qd == target:
                nyukai_today += 1
            if month_start <= qd <= target:
                nyukai_month += 1

    # --- 退塾リスト ---
    ws2 = wb["退塾リスト"]
    taikai_month = 0
    for row in range(72, 500):
        g = ws2.cell(row=row, column=7).value  # G: 年/月
        if g is None:
            continue
        gs = str(g).strip()
        if gs in (f"{year}/{month}", f"{year}/{month:02d}"):
            taikai_month += 1

    wb.close()
    return {
        "toiawase_today": toiawase_today,
        "toiawase_month": toiawase_month,
        "nyukai_today": nyukai_today,
        "nyukai_month": nyukai_month,
        "taikai_month": taikai_month,
    }


def find_thread(service, year: int, month: int) -> dict | None:
    """Gmailスレッドを検索し、最新メッセージの情報を返す"""
    query = f"subject:{CLASSROOM_NAME}日報 {year}年{month}月"
    results = service.users().messages().list(
        userId="me", q=query, maxResults=50
    ).execute()
    messages = results.get("messages", [])
    if not messages:
        return None

    # スレッドIDを取得（最初のメッセージから）
    first = service.users().messages().get(
        userId="me", id=messages[0]["id"], format="metadata",
        metadataHeaders=["Message-ID", "References", "Subject"]
    ).execute()
    thread_id = first["threadId"]

    # スレッド内の全メッセージを取得し最新を見つける
    thread = service.users().threads().get(
        userId="me", id=thread_id, format="metadata",
        metadataHeaders=["Message-ID", "References", "Subject"]
    ).execute()
    thread_msgs = thread["messages"]
    latest = thread_msgs[-1]

    headers = {h["name"]: h["value"] for h in latest["payload"]["headers"]}
    message_id = headers.get("Message-ID", "")
    references = headers.get("References", "")

    # 最新メッセージの本文と添付ファイル名を取得
    full_msg = service.users().messages().get(
        userId="me", id=latest["id"], format="full"
    ).execute()
    body_text = _extract_body_text(full_msg)

    zaiseki = _extract_number(body_text, r"月末在籍\s*(\d+)名")
    moushide = _extract_number(body_text, r"翌月申し出\s*(\d+)名")

    # 前回の添付ファイル名を取得
    prev_attachments = _extract_attachment_names(full_msg)

    return {
        "thread_id": thread_id,
        "message_id": message_id,
        "references": references,
        "zaiseki": zaiseki,
        "moushide": moushide,
        "prev_attachments": prev_attachments,
    }


def _extract_body_text(msg: dict) -> str:
    """メッセージからプレーンテキスト本文を抽出"""
    payload = msg["payload"]
    # シンプルなbody
    if "body" in payload and payload["body"].get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="ignore")
    # multipart
    for part in payload.get("parts", []):
        if part["mimeType"] == "text/plain" and part["body"].get("data"):
            return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="ignore")
        if part["mimeType"] == "text/html" and part["body"].get("data"):
            return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="ignore")
        # nested multipart
        for sub in part.get("parts", []):
            if sub.get("body", {}).get("data"):
                return base64.urlsafe_b64decode(sub["body"]["data"]).decode("utf-8", errors="ignore")
    return ""


def _extract_number(text: str, pattern: str) -> int | None:
    """正規表現で数値を抽出"""
    m = re.search(pattern, text)
    return int(m.group(1)) if m else None


def _extract_attachment_names(msg: dict) -> list[str]:
    """メッセージから添付ファイル名を抽出"""
    names = []
    payload = msg.get("payload", {})
    for part in payload.get("parts", []):
        fname = part.get("filename", "")
        if fname:
            names.append(fname)
        # nested multipart
        for sub in part.get("parts", []):
            fname = sub.get("filename", "")
            if fname:
                names.append(fname)
    return names


def find_seitobetsu_file(prev_attachments: list[str]) -> Path | None:
    """前回メールの添付から生徒別一覧ファイル名を特定し、ローカルのパスを返す"""
    for name in prev_attachments:
        if "生徒別一覧" in name:
            p = SEITOBETSU_DIR / name
            if p.exists():
                return p
    # 前回メールに見つからない場合: ディレクトリの最新
    files = sorted(SEITOBETSU_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0] if files else None


def build_html_body(target: date, counts: dict, thread_info: dict) -> str:
    """メール本文HTMLを生成"""
    m = target.month
    d = target.day
    zaiseki = thread_info.get("zaiseki")
    if zaiseki is None:
        zaiseki = "??"
    moushide = thread_info.get("moushide")
    if moushide is None:
        moushide = "??"

    return (
        f'<div dir="ltr">{m}月{d}日の日報を送付いたします。<br>'
        f"ご確認のほど宜しくお願いいたします。<br><br>"
        f"【問・入・退・在】<br>"
        f'問合せ {counts["toiawase_today"]}名（月累計{counts["toiawase_month"]}件）<br>'
        f'入会 {counts["nyukai_today"]}名（月累計{counts["nyukai_month"]}名）<br>'
        f'退会 当月 {counts["taikai_month"]}名 翌月申し出 {moushide}名<br>'
        f"月末在籍 {zaiseki}名</div>"
    )


def create_draft(service, target: date, body_html: str, thread_info: dict, attachments: list[Path]):
    """Gmail下書きを作成"""
    year, month = target.year, target.month
    subject = f"Re: {year}年{month}月 {CLASSROOM_NAME}日報"

    msg = MIMEMultipart()
    msg["To"] = NIPPOU_TO
    msg["From"] = USER_EMAIL
    msg["Subject"] = subject
    if thread_info.get("message_id"):
        msg["In-Reply-To"] = thread_info["message_id"]
        refs = thread_info.get("references", "")
        if refs:
            msg["References"] = f'{refs} {thread_info["message_id"]}'
        else:
            msg["References"] = thread_info["message_id"]

    # 署名を取得して本文に追加
    try:
        sendas = service.users().settings().sendAs().get(
            userId="me", sendAsEmail=USER_EMAIL
        ).execute()
        signature = sendas.get("signature", "")
    except Exception:
        signature = ""
    if signature:
        body_html = body_html + "<br>" + signature

    msg.attach(MIMEText(body_html, "html", "utf-8"))

    # 添付ファイル（日本語ファイル名はRFC 2231形式でエンコード）
    for fpath in attachments:
        if not fpath.exists():
            print(f"  WARN: 添付ファイルなし: {fpath}")
            continue
        with open(fpath, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        fname = fpath.name
        # RFC 2231: filename*=UTF-8''<percent-encoded>
        fname_encoded = quote(fname)
        part.add_header(
            "Content-Disposition", "attachment",
            filename=("UTF-8", "", fname),
        )
        msg.attach(part)
        print(f"  添付: {fname}")

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("ascii")

    draft_body = {"message": {"raw": raw}}
    if thread_info.get("thread_id"):
        draft_body["message"]["threadId"] = thread_info["thread_id"]

    draft = service.users().drafts().create(userId="me", body=draft_body).execute()
    return draft


def main():
    target = get_target_date()
    year, month = target.year, target.month
    print(f"対象日: {target}")

    # 1. Excel集計
    print("Excel集計中...")
    counts = count_from_excel(target)
    print(f"  問合せ: 当日{counts['toiawase_today']}名（月累計{counts['toiawase_month']}件）")
    print(f"  入会: 当日{counts['nyukai_today']}名（月累計{counts['nyukai_month']}名）")
    print(f"  退会: 当月{counts['taikai_month']}名")

    # 2. Gmail検索
    print("Gmailスレッド検索中...")
    service = get_gmail_service()
    thread_info = find_thread(service, year, month)
    if thread_info is None:
        print("  WARN: スレッドが見つかりません。新規メールとして作成します。")
        thread_info = {}
    else:
        print(f"  スレッドID: {thread_info['thread_id']}")
        print(f"  月末在籍: {thread_info.get('zaiseki', '??')}名")
        print(f"  翌月申し出: {thread_info.get('moushide', '??')}名")
        print(f"  前回添付: {thread_info.get('prev_attachments', [])}")

    # 3. 本文生成
    body_html = build_html_body(target, counts, thread_info)

    # 4. 添付ファイル（生徒別一覧は前回メールと同じものを使う）
    nippou_path = NIPPOU_DIR / "日報" / NIPPOU_EXCEL_TEMPLATE.format(year=year, month=month)
    prev_attachments = thread_info.get("prev_attachments", [])
    seitobetsu_path = find_seitobetsu_file(prev_attachments)
    attachments = [nippou_path]
    if seitobetsu_path:
        attachments.append(seitobetsu_path)
    else:
        print("  WARN: 生徒別一覧ファイルが見つかりません")

    # 5. 下書き作成
    print("下書き作成中...")
    draft = create_draft(service, target, body_html, thread_info, attachments)
    print(f"下書き作成完了: draft_id={draft['id']}")

    # 6. サマリー
    print("\n===== 日報メール =====")
    print(f"対象日: {month}月{target.day}日")
    print(f"宛先: {NIPPOU_TO}")
    print(f"件名: Re: {year}年{month}月 {CLASSROOM_NAME}日報")
    print(f"問合せ {counts['toiawase_today']}名（月累計{counts['toiawase_month']}件）")
    print(f"入会 {counts['nyukai_today']}名（月累計{counts['nyukai_month']}名）")
    print(f"退会 当月 {counts['taikai_month']}名 翌月申し出 {thread_info.get('moushide', '??')}名")
    print(f"月末在籍 {thread_info.get('zaiseki', '??')}名")
    print(f"添付: {', '.join(p.name for p in attachments)}")
    print("======================")
    print("※ 下書きのみ作成。送信はGmailから手動で行ってください。")


if __name__ == "__main__":
    main()
